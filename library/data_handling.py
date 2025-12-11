import os
import library.datetime_safety as dtms
import openpyxl as excel
import math
import cv2

class ProgramState:
    registeringstudent = False
    registeringpersonnel = False

    # The Camera
    capture = cv2.VideoCapture(1)

    def __init__(self):
        super().__init__()
        self.capture = cv2.VideoCapture(1)
    
    # For GUIs
    def change_state_student(self, state_change):
        self.registeringstudent = state_change
    
    def change_state_personnel(self, state_change):
        self.registeringpersonnel = state_change

# Checks for the accuracy of match
def face_confidence(face_distance, accuracy_threshold=0.9):
    range = (1 - accuracy_threshold)
    linear_value = (1 - face_distance) / (range * 2.0)

    if face_distance > accuracy_threshold:
        return str(round(linear_value * 100, 2)) + "%"
    else:
        value = (linear_value +((1 - linear_value) * math.pow((linear_value - 0.5 ) * 2, 0.2))) * 100
        return str(round(value, 2)) + "%"

# Error Safety incase user_data.xlsx does not exist
# Creates user_data.xlsx
# Data Structure is as follows:
# [Name of User] [Department] [Image Name Reference]
def create_user_data():
    program_dir = os.getcwd()
    object_path = os.path.join(program_dir, "user_data.xlsx")
    try:
        if os.path.exists(object_path):
            return True
        else:
            workbook = excel.Workbook()
            sheet = workbook.active

            column_titles = ["NAME", "DEPARTMENT", "IMAGE NAME REFERENCE",]

            for col_index, cell_data in enumerate(column_titles, start=1):
                sheet.cell(row=1, column=col_index, value=cell_data)
        
            sheet.title = "User Data"
            workbook.save(object_path)
            return True
    except:
        return False

# Faster method of loading user data
# Loads spreadsheet data into a dictionary of lists that only runs when the program is launched
# Data Structure is as follows:
# [Name of User] [Department] [Image Name Reference]
def load_user_data(request_state):
    program_dir = os.getcwd()
    object_path = os.path.join(program_dir, "user_data.xlsx")
    if not os.path.exists(object_path):
        create_user_data()
    workbook = excel.load_workbook(object_path)

    if request_state == "Registration":
        return object_path
    
    user_data = workbook.active
    parsed_data = {}

    for row in user_data.iter_rows(min_row=1, min_col=1, max_row=user_data.max_row+1, max_col=user_data.max_column+1, values_only=True):
        # Retrieves User Data as String
        user_name = str(row[0])
        user_department = str(row[1])
        user_filename = str(row[2])

        # User Email Deprecated
        # Could not make it work...
        # user_email = str(row[3])

        # TypeError Safety
        # Ensures all input data is a string type
        if isinstance(user_name, str) and isinstance(user_department, str) and isinstance(user_filename, str):
            parsed_data[user_filename] = [user_name, user_department, user_filename]
        else:
            print("Invalid data type in User Data")

    if not parsed_data:
        print("Parsed Data empty.")

    return parsed_data

def write_user_data(workbook, filepath, userdata):
    sheet = workbook.active
    row_number = sheet.max_row + 1

    # Locates suitable row to input data on
    for row_index, row in enumerate(sheet.iter_rows(values_only=True, min_row=1, min_col=1, max_row=sheet.max_row+1, max_col=sheet.max_column+1)):
        if row[0] == userdata[0]:
            row_number = row_index + 1
            break

    # Writes the data to said row
    for x, value in enumerate(userdata):
        if value is not None:
            sheet.cell(row=row_number, column=x+1, value=value)

    workbook.save(filepath)

# Looks for the appropriate spreadsheet to input data on
# Data Structure is as follows:
# [Name of User] [Department of User] [Time In] [Time Out]
def load_user_attendance(request_state):
    time_now = dtms.get_current_time("str_date")
    logs_path = f"attendance/{time_now}.xlsx"
    program_dir = os.getcwd()
    object_path = os.path.join(program_dir, f"{logs_path}")

    # Returns program directory path as string
    if request_state == "filepath":
        return logs_path

    # Creates a new excel workbook if no suitable workbook is found
    if not os.path.exists(object_path):
        workbook = excel.Workbook()
        sheet = workbook.active

        column_titles = ["NAME", "DEPARTMENT", "IN", "OUT"]

        for col_index, cell_data in enumerate(column_titles, start=1):
            sheet.cell(row=1, column=col_index, value=cell_data)
        
        sheet.title = time_now
        workbook.save(logs_path)
        workbook = excel.load_workbook(logs_path)
    # Loads current workbook
    else:
        workbook = excel.load_workbook(logs_path)
    
    return workbook

# Faster method of writing data into a spreadsheet
# Data Structure is as follows:
# user_data = [Name of User, Department of User]
def write_user_attendance(workbook, filepath, user_data, time_log, user_status):
    sheet = workbook.active
    row_number = sheet.max_row + 1

    # Locates suitable row to input data on
    # Checks for data in column 1 of any row
    for row_index, row in enumerate(sheet.iter_rows(values_only=True, min_row=1, min_col=1, max_row=sheet.max_row+1, max_col=sheet.max_column+1)):
        if row[0] == user_data[0]:
            row_number = row_index + 1
            break

    sheet_data_in = sheet.cell(row=row_number, column=3).value
    sheet_data_out = sheet.cell(row=row_number, column=4).value

    # Writes the data to said row
    if user_status == "IN" and sheet_data_in is None:
        data = [user_data[0], user_data[1], time_log]
        for x, value in enumerate(data):
            if value is not None:
                sheet.cell(row=row_number, column=x+1, value=value)
        workbook.save(filepath)
        return True
    elif user_status == "OUT" and sheet_data_out is None:
        data = [user_data[0], user_data[1], None, time_log]
        for x, value in enumerate(data):
            if value is not None:
                sheet.cell(row=row_number, column=x+1, value=value)
        
        workbook.save(filepath)
        return True
    else:
        return False

# Self-explanatory
# Required for GUI Buttons
def open_object(object_name):
    program_dir = os.getcwd()
    object_path = os.path.join(program_dir, f"{object_name}")
    os.startfile(object_path)

# Error Safety
# Creates folder in program directory if folder_name is not found
def directory_safety(folder_name):
    if not os.path.exists(folder_name):
        os.mkdir(folder_name)